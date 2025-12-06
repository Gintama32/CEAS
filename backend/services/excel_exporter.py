from typing import List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from db.models.data import Data


def _auto_size_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = cell.value
            try:
                length = len(str(value))
            except Exception:
                length = 0
            if length > max_length:
                max_length = length
        worksheet.column_dimensions[column].width = min(max_length + 2, 40)


def generate_estimate_workbook(rows: List[Data]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    headers = [
        "CSI Code",
        "CSI Title",
        "Description",
        "Quantity",
        "Unit",
        "Material Unit",
        "Material Amount",
        "Labor Unit",
        "Labor Amount",
        "Total Unit",
        "Total Amount",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    sorted_rows = sorted(rows, key=lambda r: (r.excel_row_number or 0))
    for row in sorted_rows:
        ws.append([
            row.csi_code or "",
            row.csi_title or "",
            row.description,
            row.quantity or "",
            row.unit or "",
            row.material_unit_cost or "",
            row.material_amount or "",
            row.labor_unit_cost or "",
            row.labor_amount or "",
            row.total_unit_cost or "",
            row.total_amount or "",
        ])

    _auto_size_columns(ws)
    return wb


def generate_raw_upload_workbook(rows: List[Data]) -> Workbook:
    """
    Create a workbook that mirrors the flat structure of the original upload
    without sections / CSI columns. Useful when a user wants the same view
    they imported, just cleaned with quantities and cost fields.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Upload"

    headers = [
        "Description",
        "Quantity",
        "Unit",
        "Material Unit Cost",
        "Material Amount",
        "Labor Unit Cost",
        "Labor Amount",
        "Total Unit Cost",
        "Total Amount",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    sorted_rows = sorted(rows, key=lambda r: (r.excel_row_number or 0))
    for row in sorted_rows:
        ws.append([
            row.description,
            row.quantity or "",
            row.unit or "",
            row.material_unit_cost or "",
            row.material_amount or "",
            row.labor_unit_cost or "",
            row.labor_amount or "",
            row.total_unit_cost or "",
            row.total_amount or "",
        ])

    _auto_size_columns(ws)
    return wb

