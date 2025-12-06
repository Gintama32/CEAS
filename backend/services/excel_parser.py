import re
from io import BytesIO
from typing import List, Dict, Tuple, Optional, Any

import openpyxl

GENERIC_COLUMN_PATTERN = re.compile(r'^column[\s_]?(\d+)$', re.IGNORECASE)
CSI_HEADER_PATTERN = re.compile(
    r'^(?P<first>\d{2})(?:[-\s]?)(?P<middle>\d{2})(?:[-\s]?)(?P<last>\d{2})(?:\s+|[-\s]+)?(?P<title>.*)$'
)


class ExcelParsingError(Exception):
    """Raised when the uploaded Excel file cannot be parsed."""


def parse_estimate_rows(
    file_bytes: bytes,
    sheet_name: Optional[str] = "EST SITE",
    header_main_row: int = 10,
    header_sub_row: int = 11,
    data_start_row: int = 12,
) -> List[Dict[str, Any]]:
    """
    Parse the estimate worksheet and return structured row dictionaries.
    """
    if not file_bytes:
        return []

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ExcelParsingError("Unable to open Excel workbook.") from exc

    target_sheet_name = sheet_name if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[target_sheet_name]
    rows = list(sheet.iter_rows(values_only=True))

    detected_header_main, detected_header_sub, detected_data_start = _detect_header_rows(
        rows,
        search_limit=25,
    )

    effective_header_main = detected_header_main or header_main_row
    effective_header_sub = detected_header_sub or header_sub_row
    effective_data_start = detected_data_start or data_start_row

    return _extract_records_from_rows(
        rows,
        header_main_row=effective_header_main,
        header_sub_row=effective_header_sub,
        data_start_row=effective_data_start,
    )


def _detect_header_rows(
    rows,
    search_limit: int = 25,
) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Attempt to auto-detect header rows by searching for a row containing "Description".
    Returns (header_main_row, header_sub_row, data_start_row) if detected.
    """
    search_limit = min(search_limit, len(rows))

    for idx in range(search_limit):
        row = rows[idx]
        if not row:
            continue

        if any(isinstance(cell, str) and cell.strip().lower() == "description" for cell in row):
            header_main = idx + 1
            header_sub = header_main + 1 if idx + 1 < len(rows) else header_main
            data_start = header_sub + 1 if header_sub < len(rows) else header_sub
            return header_main, header_sub, data_start

    return None, None, None


def _extract_records_from_rows(
    rows,
    header_main_row: int,
    header_sub_row: int,
    data_start_row: int,
) -> List[Dict[str, Any]]:
    if not rows:
        return []

    header_row_1 = rows[header_main_row - 1] if len(rows) >= header_main_row else []
    header_row_2 = rows[header_sub_row - 1] if len(rows) >= header_sub_row else []
    headers = _combine_headers(header_row_1, header_row_2)

    meaningful_indices = [
        idx for idx, header in enumerate(headers) if _is_meaningful_header(header)
    ]

    data_rows = rows[data_start_row - 1 :]

    records: List[Dict[str, Any]] = []
    current_csi_code: Optional[str] = None
    current_csi_title: Optional[str] = None

    for row_offset, row in enumerate(data_rows):
        excel_row_number = data_start_row + row_offset
        row_values = _collect_row_values(row, headers, meaningful_indices)

        description_raw = _get_first_value(row_values, "Description")
        description_text = _clean_string(description_raw)

        has_other_values = any(
            _has_meaningful_value(value)
            for header, value in row_values
            if header != "Description"
        )

        if description_text and not has_other_values:
            header_match = CSI_HEADER_PATTERN.match(description_text)
            if header_match:
                formatted_code = _format_csi_code(
                    header_match.group('first'),
                    header_match.group('middle'),
                    header_match.group('last'),
                )
                current_csi_code = formatted_code
                title = header_match.group('title').strip() if header_match.group('title') else None
                current_csi_title = title or None
            # Skip header or divider rows regardless of match outcome
            continue

        if _is_divider_row(row_values):
            continue

        if _is_total_row(description_text, row_values):
            continue

        if not any(_has_meaningful_value(value) for _, value in row_values):
            continue

        if not description_text:
            # Skip rows without description (usually totals or spacing rows)
            continue

        record = {
            "section": None,
            "subsection": None,
            "csi_code": current_csi_code,
            "csi_title": current_csi_title,
            "description": description_text,
            "excel_row_number": excel_row_number,
            "quantity": _to_float(_get_first_value(row_values, "Quantity")),
            "unit": _clean_string(_get_first_value(row_values, "Unit")),
            "material_unit_cost": _to_float(_get_value_by_prefix(row_values, "Material - Unit")),
            "material_amount": _to_float(_get_value_by_prefix(row_values, "Material - Amount")),
            "labor_unit_cost": _to_float(_get_value_by_prefix(row_values, "Labor - Unit")),
            "labor_amount": _to_float(_get_value_by_prefix(row_values, "Labor - Amount")),
            "total_unit_cost": _to_float(_get_value_by_prefix(row_values, "Total - Unit")),
            "total_amount": _to_float(_get_value_by_prefix(row_values, "Total - Amount")),
        }

        records.append(record)

    return records


def _format_csi_code(first: str, middle: str, last: str) -> str:
    return f"{first} {middle} {last}"


def _combine_headers(row1, row2) -> List[Optional[str]]:
    headers: List[Optional[str]] = []
    max_cols = max(len(row1 or []), len(row2 or []))
    current_main_header: Optional[str] = None
    header_counts: Dict[str, int] = {}

    for idx in range(max_cols):
        header_1 = row1[idx] if row1 and idx < len(row1) else None
        header_2 = row2[idx] if row2 and idx < len(row2) else None

        main_header = _clean_string(header_1)
        sub_header = _clean_string(header_2)

        if main_header:
            current_main_header = main_header

        header_name: Optional[str] = None
        if sub_header and current_main_header:
            header_name = f"{current_main_header} - {sub_header}"
        elif sub_header:
            header_name = sub_header
        elif main_header:
            header_name = main_header

        if header_name:
            count = header_counts.get(header_name, 0) + 1
            header_counts[header_name] = count
            if count > 1:
                header_name = f"{header_name} [{count}]"
            headers.append(header_name)
        else:
            headers.append(None)

    return headers


def _is_meaningful_header(header: Optional[str]) -> bool:
    if not header:
        return False
    return GENERIC_COLUMN_PATTERN.match(header.strip().lower()) is None


def _collect_row_values(row, headers, indices) -> List[Tuple[Optional[str], Any]]:
    values: List[Tuple[Optional[str], Any]] = []
    for idx in indices:
        header = headers[idx]
        value = row[idx] if row and idx < len(row) else None
        values.append((header, value))
    return values


def _get_first_value(row_values: List[Tuple[Optional[str], Any]], target: str) -> Any:
    for header, value in row_values:
        if header == target:
            return value
    return None


def _normalize_header(header: Optional[str]) -> Optional[str]:
    if not header:
        return None
    return header.split(" [", 1)[0]


def _get_value_by_prefix(row_values: List[Tuple[Optional[str], Any]], target_prefix: str) -> Any:
    for header, value in row_values:
        normalized = _normalize_header(header)
        if normalized == target_prefix:
            return value
    return None


def _clean_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    return text


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text == "-":
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def _has_meaningful_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip()
    if not text or text == "-":
        return False
    if re.fullmatch(r"-+", text):
        return False
    return True


def _is_divider_row(row_values: List[Tuple[Optional[str], Any]]) -> bool:
    return any(
        isinstance(value, str) and value.strip() and re.fullmatch(r"-+", value.strip())
        for _, value in row_values
    )


def _is_total_row(description_text: Optional[str], row_values: List[Tuple[Optional[str], Any]]) -> bool:
    desc_lower = description_text.lower() if description_text else ""
    if not description_text:
        has_total_value = any(
            (header and header.lower().startswith("total")) and _has_meaningful_value(value)
            for header, value in row_values
        )
        if has_total_value:
            has_non_total_value = any(
                (header and not header.lower().startswith("total")) and _has_meaningful_value(value)
                for header, value in row_values
            )
            if not has_non_total_value:
                return True
    if "grand total" in desc_lower:
        return True
    if desc_lower.startswith("total") and not desc_lower.startswith("subtotal"):
        return True
    return False

